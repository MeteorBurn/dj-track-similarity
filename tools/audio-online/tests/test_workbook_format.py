from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from metadata_enrichment.workbook_format import format_workbook


def test_formatter_expands_each_column_and_disables_wrapping(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Track Name", "MAEST Genres"])
    sheet.append(["Artist — A very long title that must stay in one cell", "Techno (52%); Electro (38%); House (36%)"])
    workbook.save(path)

    format_workbook(path)

    formatted = load_workbook(path)
    sheet = formatted.active
    assert sheet.column_dimensions["A"].width >= len(str(sheet["A2"].value)) + 2
    assert sheet.column_dimensions["B"].width >= len(str(sheet["B2"].value)) + 2
    assert sheet["A2"].alignment.wrap_text is not True
