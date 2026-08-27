"""Final XLSX sizing with openpyxl after the artifact-tool export."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def format_workbook(path: Path) -> None:
    """Fit every populated column to its longest value and prohibit text wrapping."""

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for column_index in range(1, sheet.max_column + 1):
            width = 0
            for row_index in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                alignment = copy(cell.alignment)
                alignment.wrap_text = False
                cell.alignment = alignment
                width = max(width, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(column_index)].width = width + 2
    workbook.save(path)
